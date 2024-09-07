from new import add_data
from openpyxl import load_workbook
from trial import send_mail

while True:

    print("""1. Computer intelligence
  2. Python
  3. Data mining""")

    n = int(input("Enter the subject number: "))
    no_of_absent = int(input("Enter the number of absentees: "))
    lst_of_absent = []

    if no_of_absent == 1:
        lst_of_absent.append(int(input("Enter the roll no of the absentee: ")))

    else:
        for i in range(no_of_absent):
            lst_of_absent.append(
                int(input(f"Enter the roll no of absentee no.{i+1}: ")))

    column_name = chr(66 + n)

    wb = load_workbook(r"C:\Users\kavit\OneDrive\Desktop\Attendance.xlsx")
    ws = wb.active

    for absent in lst_of_absent:
        ws[column_name + str(absent+1)].value += 1
        if ws[column_name + str(absent+1)].value == 2:
            send_mail(1, [absent,  ws["B" + str(absent+1)].value],
                      ws[column_name + str(1)].value)
        elif ws[column_name + str(absent+1)].value == 3:
            send_mail(2, [absent,  ws["B" + str(absent+1)].value],
                      ws[column_name + str(1)].value)

    wb.save(r"C:\Users\kavit\OneDrive\Desktop\Attendance.xlsx")

    cont = int(input("Another subject? (1 for yes, 0 for no): "))
    if cont == 0:
        break
